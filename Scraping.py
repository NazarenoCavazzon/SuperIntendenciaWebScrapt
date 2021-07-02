from openpyxl import Workbook, load_workbook
import os

def clear(): return os.system("cls")

def writeOnExcel():
    from openpyxl import Workbook, load_workbook

def get_columns(number):
    lista=[]
    nu = 0
    import requests
    from bs4 import BeautifulSoup as Soup

    url = "https://www.ssn.gob.ar/storage/registros/productores/productoresactivos.asp"

    payload = {
        "socpro": "PAS",
        "matricula": "{}".format(number),
        "apellidorazonsocial": "",
        "docNro": "",
        "Submit": "Buscar"
    }

    response = requests.post(url, data=payload)
    response.raise_for_status()

    soup = Soup(response.content, "html.parser")

    for column in soup.select("div[class^=\"col-md-\"]"):
        if nu < 13 and nu >= 2:
            lista.append(" ".join(column.get_text().strip().split()))
            nu +=1
        elif nu < 2:
            nu += 1
    for i,j in enumerate(lista):
        valor = []
        valor.append(j.split(":"))
        if len(valor[0]) == 2:
            lista[i] = valor[0][1]

    return lista

#Varios sin nada

def main():
    porcentaje = 0
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Matricula","Nombre","DNI","Cuit","Ramo","Domicilio","Localidad","Provincia","Cod. Postal", "Telefono", "Mail"])
    for i in range(0,100000):
        listota = []
        listota = get_columns(i)
        if len(listota) != 0:
            ws.append(listota)
        porcentaje = round(((i/100000)*100),1)
        clear()
        print(f"{porcentaje}%")
    wb.save("Final Result.xlsx")

if __name__ == "__main__":
    main()